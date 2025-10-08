#!/usr/bin/env python
# Copyright 2021 - 2025 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Script to generate submission Excel sheet"""

import glob
import os
import sys
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Final, Generator, Mapping

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import BORDER_THIN, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field
from schemapack import load_schemapack
from schemapack.spec.schemapack import ClassDefinition, SchemaPack
from script_utils.cli import echo_failure, echo_success, run
from typer import Option

HERE: Final[Path] = Path(__file__).parent.resolve()
SCHEMAPACK_PATH: Final[Path] = (
    HERE.parent / "build" / "ghga_metadata_schema.resolved.schemapack.yaml"
)
CONFIG_PATH: Final[Path] = HERE.parent / "spreadsheet_conf.yaml"
XLSX_DIR: Final[Path] = HERE.parent / "spreadsheets"

THIN_SIDE = Side(border_style=BORDER_THIN, color="00000000")
THIN_GRAY_SIDE = Side(border_style=BORDER_THIN, color="808080")

THIN_BORDER: Final[Border] = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
THIN_BORDER_GRAY: Final[Border] = Border(
    left=THIN_GRAY_SIDE,
    right=THIN_GRAY_SIDE,
    top=THIN_GRAY_SIDE,
    bottom=THIN_GRAY_SIDE,
)
ALIGN_HEADER: Final[Alignment] = Alignment(
    wrapText=True, horizontal="center", vertical="top"
)


class GHGASchemaError(RuntimeError):
    """Raised when a schema violates GHGA schema rules."""


class XLSXContentDifference(RuntimeError):
    """Raised when a content difference was detected between the expected and
    observed XLSX files.
    """


class WorksheetStyle(BaseModel):
    """Style configuration for a worksheet"""

    header_color: str | None = None
    content_color: str | None = None


class Config(BaseModel):
    """A XLSX generator config"""

    output_filename: str
    styles: dict[str, WorksheetStyle] = Field(default_factory=dict)


@dataclass
class IdAnnotatedClass:
    """
    Describes a class name and its associated ID property name.

    Attributes:
        class_name (str): The name of the class.
        id_property_name (str): The name of the property that serves as the unique identifier for the class.
    """

    class_name: str
    id_property_name: str


@dataclass
class Column:
    """Represents metadata for a column in a sheet.

    Attributes:
        name: The name of the column.
        description: An optional description explaining the column's meaning.
        type: The data type of the column's values (e.g., "string", "int").
        multivalued: Whether the column can contain multiple values per cell.
        class_reference: A reference to another class if the column links to an external entity.
        enum: Whether the column is restricted to values from a controlled vocabulary.
        required: Whether the column must have a value in every row.

    """

    name: str
    description: str | None
    type: str
    multivalued: bool
    class_reference: IdAnnotatedClass | None
    is_enum: bool
    required: bool

    @property
    def restriction(self) -> str:
        if self.is_enum and self.class_reference:
            raise GHGASchemaError("Column cannot be both an enum and a reference")
        # If the column is an enum, its values are restricted to a predefined set (controlled vocabulary).
        if self.is_enum:
            return "restriction: value from controlled vocabulary (see schema for allowed values)"
        if self.class_reference:
            return f"restriction: value from {self.class_reference.class_name}.{self.class_reference.id_property_name}"
        return "unrestricted"


@dataclass
class WorksheetMetadata:
    """
    Metadata defining the structure and styling of a sheet.

    Attributes:
        columns (list[Column]): The columns present in the worksheet.
        id_property_name (str): The name of the property that serves as the unique identifier for the worksheet's class.
        header_color (None | str): The color used for the header row, if any.
        content_color (None | str): The color used for the content rows, if any.
    """

    class_name: str
    columns: list[Column]
    id_property_name: str
    header_color: None | str
    content_color: None | str


def _get_element_type_from_schema(schema: dict) -> str:
    """Get the type of the elements in an array from a JSON schema."""
    element_type = schema.get("type", "object")
    if isinstance(element_type, tuple):
        element_type = ", ".join(str(t) for t in element_type)
    if "array" in element_type:
        return schema.get("items", {}).get("type", "object")
    return element_type


class ClassColumnFactory:
    """Factory for creating columns from class definitions"""

    @staticmethod
    def create_id_column(class_definition: ClassDefinition) -> Column:
        """Creates a column representing the ID property of a class definition."""
        return Column(
            name=class_definition.id.propertyName,
            description=class_definition.id.description,
            type="string",
            is_enum=False,
            multivalued=False,
            class_reference=None,
            required=True,
        )

    @staticmethod
    def create_content_columns(class_definition: ClassDefinition) -> list[Column]:
        """Extracts columns from the content schema of a class definition. Returns a list of
        columns representing the properties defined in the class's content schema."""

        properties = class_definition.content.get("properties", {})
        required_fields = class_definition.content.get("required", [])

        return [
            Column(
                name=property_name,
                description=properties[property_name].get("description", ""),
                type=_get_element_type_from_schema(properties[property_name]),
                multivalued="array" in properties[property_name]["type"],
                class_reference=None,
                is_enum="enum" in properties[property_name],
                required=property_name in required_fields,
            )
            for property_name in properties
        ]

    @staticmethod
    def create_relations_columns(
        class_definition: ClassDefinition, model: SchemaPack
    ) -> list[Column]:
        """Extracts columns from the relations of a class definition."""

        class_relations = class_definition.relations
        return [
            Column(
                name=relation_property_name,
                description=class_relation.description,
                type="string",
                is_enum=False,
                multivalued=class_relation.multiple.target,
                class_reference=IdAnnotatedClass(
                    class_name=class_relation.targetClass,
                    id_property_name=model.classes[
                        class_relation.targetClass
                    ].id.propertyName,
                ),
                required=class_relation.mandatory.origin,
            )
            for relation_property_name, class_relation in class_relations.items()
        ]


class FormatUtils:
    """Utility class for formatting worksheets."""

    def __init__(self, worksheet: Worksheet) -> None:
        self.worksheet = worksheet

    def _make_value_cell(self, fill_content: PatternFill | None) -> Cell:
        """Creates a new cell for the value area of the worksheet"""
        value_cell = Cell(self.worksheet)  # type: ignore
        if fill_content:
            value_cell.fill = fill_content
        value_cell.border = THIN_BORDER_GRAY
        return value_cell

    def _make_header_bold(self) -> None:
        """Makes the first row bold."""
        first_row = next(self.worksheet.iter_rows(min_row=1, max_row=1))
        for cell in first_row:
            cell.font = Font(bold=True)

    def _apply_cell_alignment_and_borders(self, fill_header: PatternFill | None) -> None:
        """Aligns and borders all cells in the worksheet."""
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.alignment = ALIGN_HEADER
                if fill_header:
                    cell.fill = fill_header
                cell.border = THIN_BORDER

    def _color_header(self, header_color: str | None) -> None:
        """Colors the header row of the worksheet."""
        if header_color:
            self.worksheet.sheet_properties.tabColor = header_color

    def _fill_with_value_cells(
        self, fill_content: PatternFill | None, column_number: int
    ) -> None:
        """Appends 1000 rows of value cells to the worksheet."""
        for _ in range(1000):
            self.worksheet.append(
                [self._make_value_cell(fill_content) for _ in range(column_number)]
            )

    def _set_column_widths(self) -> None:
        """Sets the width of all columns to 35."""
        for column in range(1, self.worksheet.max_column + 1):
            self.worksheet.column_dimensions[get_column_letter(column)].width = 35


def generate_worksheet_metadata(
    class_name: str,
    class_definition: ClassDefinition,
    model: SchemaPack,
    config: Config,
) -> WorksheetMetadata:
    """Builds a worksheet from a class definition. Combined the column specs with
    the worksheet style parameters.
    """
    columns: list[Column] = []

    # ID column
    columns.append(ClassColumnFactory.create_id_column(class_definition))

    # Relations
    columns.extend(ClassColumnFactory.create_relations_columns(class_definition, model))

    # Content
    columns.extend(ClassColumnFactory.create_content_columns(class_definition))

    style = config.styles.get(class_name, WorksheetStyle())
    return WorksheetMetadata(
        class_name=class_name,
        columns=columns,
        id_property_name=class_definition.id.propertyName,
        header_color=style.header_color,
        content_color=style.content_color,
    )


def annotate_worksheet(
    worksheet_metadata: WorksheetMetadata, worksheet: Worksheet
) -> Worksheet:
    """
    Annotates a worksheet object from openpyxl with metadata from WorksheetMetadata.
    """
    def get_column_spec(col):
        return [
            col.name,
            col.description,
            col.type,
            "multiple values" if col.multivalued else "single value",
            col.restriction,
            "required" if col.required else "optional",
        ]

    for col in worksheet_metadata.columns:
        worksheet.append(get_column_spec(col))

    return worksheet


def format_worksheet(worksheet_metadata: WorksheetMetadata, worksheet: Worksheet):
    """Formats a worksheet by defining the styles of the cells."""

    header_color = worksheet_metadata.header_color
    content_color = worksheet_metadata.content_color
    fill_header = PatternFill("solid", fgColor=header_color) if header_color else None
    fill_content = (
        PatternFill("solid", fgColor=content_color) if content_color else None
    )
    format_worksheet = FormatUtils(worksheet)

    format_worksheet._make_header_bold()
    format_worksheet._apply_cell_alignment_and_borders(fill_header)
    format_worksheet._color_header(header_color)
    format_worksheet._fill_with_value_cells(
        fill_content, len(worksheet_metadata.columns)
    )
    format_worksheet._set_column_widths()

    return worksheet


def generate_worksheet_metadata_for_all_classes(
    model: SchemaPack, config: Config
) -> Mapping[str, WorksheetMetadata]:
    """Generates worksheet metadata for all classes in the schema pack."""
    return {
        class_name: generate_worksheet_metadata(
            class_name, class_definition, model, config
        )
        for class_name, class_definition in model.classes.items()
    }


def generate_workbook(sheets: Mapping[str, WorksheetMetadata]) -> Workbook:
    """Generates a workbook from a dictionary of sheets."""
    # Create a new workbook
    workbook = Workbook()
    # Remove the default sheet
    workbook.remove(workbook.worksheets[0])

    for sheet_name, sheet_metadata in sheets.items():
        # Create a new sheet for the class
        workbook_sheet = workbook.create_sheet(title=sheet_name)

        workbook_sheet = annotate_worksheet(sheet_metadata, workbook_sheet)
        workbook_sheet = format_worksheet(sheet_metadata, workbook_sheet)

    return workbook


@contextmanager
def working_directory(working_dir: Path) -> Generator[None, None, None]:
    """Temporarily changes the working directory to the specified directory"""
    cwd = Path.cwd()
    try:
        os.chdir(working_dir)
        yield
    finally:
        os.chdir(cwd)


def compare_xls(expected: Workbook, observed: Workbook) -> None:
    """
    Compares two openpyxl Workbook objects for differences.

    Parameters:
        expected (Workbook): The expected workbook to compare against.
        observed (Workbook): The observed workbook to check for differences.

    Raises:
        XLSXContentDifference: If any differences are detected between the workbooks.

    """
    if expected.sheetnames != observed.sheetnames:
        raise XLSXContentDifference(
            f"Sheets differ. expected={expected.sheetnames}. observed={observed.sheetnames}"
        )
    cell_style_attributes = ["alignment", "font", "border", "fill"]

    for sheet in expected.sheetnames:
        for row_idx, (row_a, row_b) in enumerate(
            zip(expected[sheet].iter_rows(), observed[sheet].iter_rows()), 1
        ):
            for col_idx, (cell_a, cell_b) in enumerate(zip(row_a, row_b), 1):
                if cell_a.value != cell_b.value:
                    raise XLSXContentDifference(
                        f"Cell values differ in sheet {sheet}, row {row_idx}, col {col_idx}: "
                        + f"expected={cell_a.value}, observed={cell_b.value}"
                    )
                for attr in cell_style_attributes:
                    style_a = getattr(cell_a, attr)
                    style_b = getattr(cell_b, attr)
                    if style_a != style_b:
                        raise XLSXContentDifference(
                            f"Cell {attr} differs in sheet {sheet}, row {row_idx}, col {col_idx}: "
                            + f"expected={style_a}, "
                            + f"observed={style_b}"
                        )


def compare_folders(expected: Path, observed: Path):
    """Function to check equality of contents of two folders with xlsx files."""

    with working_directory(expected):
        expected_glob = glob.glob("*")

    with working_directory(observed):
        observed_glob = glob.glob("*")

    if sorted(expected_glob) != sorted(observed_glob):
        raise XLSXContentDifference(
            f"Different directory contents. expected={expected_glob}. observed={observed_glob}"
        )

    for fname in expected_glob:
        expected_wb = load_workbook(expected.joinpath(fname))
        observed_wb = load_workbook(observed.joinpath(fname))
        try:
            compare_xls(expected=expected_wb, observed=observed_wb)
        except XLSXContentDifference as err:
            raise XLSXContentDifference(f"{fname}: {err}")


HEADER_ROW_INDEX = 1
DATA_START_ROW_INDEX = 7
META_SHEET_HEADER = ["sheet", "n_cols", "header_row", "data_start", "id_property_name"]
COLUMN_META_SHEET_HEADER = [
    "sheet",
    "column",
    "multivalued",
    "type",
    "ref_class",
    "ref_class_id_property",
    "enum",
    "required",
]


def _add_version_sheet(model: SchemaPack, workbook: Workbook) -> None:
    """Adds a model version sheet to the workbook"""
    version_sheet = workbook.create_sheet("__version")
    version_sheet.sheet_state = "hidden"
    version_sheet.append([model.schemapack])


def _add_meta_sheet(workbook: Workbook) -> None:
    """Adds a meta sheet to the workbook and adds a header that corresponds to the
    worksheet's name, number of columns, header row number, and data row start number.
    """
    meta_sheet = workbook.create_sheet("__sheet_meta")
    meta_sheet.sheet_state = "hidden"
    meta_sheet.append(META_SHEET_HEADER)


def _populate_meta_sheet(
    meta_sheet: Worksheet, meta_sheet_content: WorksheetMetadata
) -> None:
    """Populates the meta sheet with the information of a given worksheet metadata."""

    meta_sheet.append(
        [
            meta_sheet_content.class_name,
            len(meta_sheet_content.columns),
            HEADER_ROW_INDEX,
            DATA_START_ROW_INDEX,
            meta_sheet_content.id_property_name,
        ]
    )


def _add_column_meta_sheet(workbook: Workbook) -> None:
    """Adds a column meta sheet to the workbook and adds a header that corresponds to the"""
    column_meta_sheet = workbook.create_sheet("__column_meta")
    column_meta_sheet.sheet_state = "hidden"
    column_meta_sheet.append(COLUMN_META_SHEET_HEADER)


def _populate_column_meta_sheet(
    column_meta_sheet: Worksheet, meta_sheet_content: WorksheetMetadata
) -> None:
    """Populates the column meta sheet with the information of a given worksheet metadata."""

    for column in meta_sheet_content.columns:
        column_meta_sheet.append(
            [
                meta_sheet_content.class_name,
                column.name,
                column.multivalued,
                column.type,
                column.class_reference.class_name if column.class_reference else None,
                column.class_reference.id_property_name
                if column.class_reference
                else None,
                column.is_enum,
                column.required,
            ]
        )


def add_transpiler_metadata(
    model: SchemaPack,
    workbook: Workbook,
    sheets_metadata: Mapping[str, WorksheetMetadata],
) -> None:
    """Adds metadata required for parsing it with ghga-transpiler to the workbook"""

    _add_version_sheet(model, workbook)
    _add_meta_sheet(workbook)
    _add_column_meta_sheet(workbook)
    for _, worksheet_metadata in sheets_metadata.items():
        _populate_meta_sheet(workbook["__sheet_meta"], worksheet_metadata)
        _populate_column_meta_sheet(workbook["__column_meta"], worksheet_metadata)


def load_config(config_path: Path) -> Config:
    """Loads a config from a file."""
    with config_path.open("r") as file:
        return Config.model_validate(yaml.safe_load(file))


def create_xlsx_files(xlsx_dir: Path = XLSX_DIR) -> None:
    """The logic for generating a workbook. Loads the schemapack and config,
    generates the workbook, and saves it to a file.
    """
    schema = load_schemapack(Path(SCHEMAPACK_PATH))
    config = load_config(Path(CONFIG_PATH))
    sheets_metadata = generate_worksheet_metadata_for_all_classes(schema, config)
    workbook = generate_workbook(sheets_metadata)
    add_transpiler_metadata(schema, workbook, sheets_metadata)
    workbook.save(xlsx_dir / config.output_filename)


def main(
    check: bool = Option(
        False,
        help="Run read-only check that verifies that the documents are up-to-date.",
    ),
):
    """Generate a submission XLSX file from the schema pack."""
    if check:
        with TemporaryDirectory() as tmpdirname:
            tmp_docs_dir = Path(tmpdirname)

            create_xlsx_files(xlsx_dir=tmp_docs_dir)

            try:
                compare_folders(XLSX_DIR, tmp_docs_dir)
                echo_success("Documents are up-to-date")
            except XLSXContentDifference as err:
                echo_failure("Documents are not up-to-date")
                echo_failure(str(err))
                sys.exit(1)
    else:
        create_xlsx_files()


if __name__ == "__main__":
    run(main)
